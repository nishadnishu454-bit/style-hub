from django.shortcuts import render
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from user.orders.models import Order, OrderItem
from admin_panel.productmanagement.models import Product
from django.core.paginator import Paginator
from user.authentication.models import Referral
from admin_panel.categorymanagement.models import Category

User = get_user_model()


def is_admin(user):
    return user.is_authenticated and user.is_staff


def get_filtered_orders(filter_type, start_date_str=None, end_date_str=None):
    orders = Order.objects.all()    
    today = timezone.localdate()
    
    if filter_type == 'daily':
        orders = orders.filter(ordered_at__date=today)
    elif filter_type == 'weekly':
        start_date = today - timedelta(days=6)
        orders = orders.filter(ordered_at__date__gte=start_date, ordered_at__date__lte=today)
    elif filter_type == 'monthly':
        start_date = today - timedelta(days=29)
        orders = orders.filter(ordered_at__date__gte=start_date, ordered_at__date__lte=today)
    elif filter_type == 'yearly':
        orders = orders.filter(ordered_at__year=today.year)
    elif filter_type == 'custom':
        if start_date_str:
            try:
                s_date = datetime.strptime(start_date_str.strip(), '%Y-%m-%d').date()
                orders = orders.filter(ordered_at__date__gte=s_date)
            except ValueError:
                pass
        if end_date_str:
            try:
                e_date = datetime.strptime(end_date_str.strip(), '%Y-%m-%d').date()
                orders = orders.filter(ordered_at__date__lte=e_date)
            except ValueError:
                pass
    return orders


@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='admin_login')
def admin_dashboard(request):
    chart_filter = request.GET.get('chart_filter', 'monthly')  # 'weekly', 'monthly', 'yearly'
    
    total_users = User.objects.filter(is_staff=False).count()
    active_users = User.objects.filter(is_staff=False, is_active=True).count()
    blocked_users = User.objects.filter(is_staff=False, is_active=False).count()
    total_products = Product.objects.filter(is_deleted=False).count()
    total_categories = Category.objects.filter(is_deleted=False).count()
    
    active_orders = Order.objects.exclude(order_status__in=['Cancelled', 'Returned'])
    total_orders = active_orders.count()
    total_revenue_val = active_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    
    pending_orders = Order.objects.filter(order_status='Pending').count()
    confirmed_orders = Order.objects.filter(order_status='Confirmed').count()
    shipped_orders = Order.objects.filter(order_status='Shipped').count()
    out_for_delivery_orders = Order.objects.filter(order_status='Out for Delivery').count()
    delivered_orders = Order.objects.filter(order_status='Delivered').count()
    cancelled_orders = Order.objects.filter(order_status='Cancelled').count()
    
    full_returns = Order.objects.filter(order_status='Return Requested').count()
    partial_returns = OrderItem.objects.filter(item_status='Return Requested').exclude(order__order_status='Return Requested').values('order').distinct().count()
    return_orders = full_returns + partial_returns
    returned_orders = Order.objects.filter(order_status='Returned').count()
    
    recent_orders = Order.objects.all().select_related('user').order_by('-ordered_at')[:5]
    


    sales_chart = []
    today = timezone.localdate()
    
    if chart_filter == 'weekly':
        # Last 7 days
        days = [today - timedelta(days=i) for i in range(6, -1, -1)]
        raw_chart_data = []
        for d in days:
            amt = Order.objects.exclude(order_status__in=['Cancelled', 'Returned']).filter(
                ordered_at__date=d
            ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
            raw_chart_data.append({'label': d.strftime('%a'), 'amount': float(amt)})
            
    elif chart_filter == 'yearly':
        raw_chart_data = []
        for m in range(1, 13):
            amt = Order.objects.exclude(order_status__in=['Cancelled', 'Returned']).filter(
                ordered_at__year=today.year,
                ordered_at__month=m
            ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
            month_name = datetime(today.year, m, 1).strftime('%b')
            raw_chart_data.append({'label': month_name, 'amount': float(amt)})
            
    else:  
        days = [today - timedelta(days=i) for i in range(14, -1, -1)]
        raw_chart_data = []
        for d in days:
            amt = Order.objects.exclude(order_status__in=['Cancelled', 'Returned']).filter(
                ordered_at__date=d
            ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
            raw_chart_data.append({'label': d.strftime('%d %b'), 'amount': float(amt)})
            

    max_amount = max([item['amount'] for item in raw_chart_data]) if raw_chart_data else 0
    for item in raw_chart_data:
        height = int((item['amount'] / max_amount) * 100) if max_amount > 0 else 0
        if item['amount'] > 0 and height < 5:
            height = 5
        sales_chart.append({
            'label': item['label'],
            'amount': item['amount'],
            'height': height
        })
        

    top_products = OrderItem.objects.exclude(order__order_status__in=['Cancelled', 'Returned']
        ).exclude( item_status__in=['Cancelled', 'Returned', 'Return Requested']
        ).values('product_name','variant__product_id'
        ).annotate(
            total_qty=Sum('quantity'),
            total_sales=Sum('total_price')
        ).order_by('-total_qty')[:10]
        

    top_categories = OrderItem.objects.exclude(order__order_status__in=['Cancelled', 'Returned']).exclude(
            item_status__in=['Cancelled', 'Returned', 'Return Requested']
        ).values( 'variant__product__category__category_name'
        ).annotate(
            total_qty=Sum('quantity'),
            total_sales=Sum('total_price')
        ).order_by('-total_qty')[:10]
        

        
    context = {
        'total_revenue': total_revenue_val,
        'total_orders': total_orders,
        'total_users': total_users,
        'active_users': active_users,
        'blocked_users': blocked_users,
        'total_products': total_products,
        'total_categories':total_categories,
        'pending_orders': pending_orders,
        'confirmed_orders': confirmed_orders,
        'shipped_orders': shipped_orders,
        'out_for_delivery_orders': out_for_delivery_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        'return_orders': return_orders,
        'returned_orders':returned_orders,
        'recent_orders': recent_orders,
        'sales_chart': sales_chart,
        'chart_filter': chart_filter,
        'top_products': top_products,
        'top_categories': top_categories,
    }
    return render(request, 'admindashboard/admin_dashboard.html', context)


@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='admin_login')
def sales_report_page(request):

    filter_type = request.GET.get('filter_type', 'daily')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    
    all_orders = get_filtered_orders(filter_type, start_date, end_date)
    orders = all_orders.exclude(order_status__in=['Cancelled', 'Returned'])
    refunded_orders = all_orders.filter(payment_status='Refunded')
    refund_amount = refunded_orders.aggregate(total=Sum('refunded_amount'))['total'] or Decimal('0.00')

    # MAIN STATS
    stats = orders.aggregate(
        total_orders_count=Count('id'),
        total_subtotal_sum=Sum('subtotal'),
        total_discount_sum=Sum('discount_amount'),
        total_delivery_sum=Sum('delivery_charge'),
        total_amount_sum=Sum('total_amount')
    )

    total_orders_count = stats['total_orders_count'] or 0
    total_subtotal = stats['total_subtotal_sum'] or Decimal('0.00')
    total_discount = stats['total_discount_sum'] or Decimal('0.00')
    total_delivery = stats['total_delivery_sum'] or Decimal('0.00')
    gross_revenue = stats['total_amount_sum'] or Decimal('0.00')

    
    net_revenue = gross_revenue - refund_amount

    # PAGINATION
    paginator = Paginator(all_orders.order_by('-ordered_at'), 15)
    page_number = request.GET.get('page')
    orders_page = paginator.get_page(page_number)

    context = {
        'orders': orders_page,
        'filter_type': filter_type,
        'start_date': start_date,
        'end_date': end_date,
        'total_orders_count': total_orders_count,
        'total_subtotal': total_subtotal,
        'total_discount': total_discount,
        'total_delivery': total_delivery,
        'refund_amount': refund_amount,
        'net_revenue': net_revenue,
    }

    return render(request,'admindashboard/sales_report.html', context)



@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='admin_login')
def download_sales_report_pdf(request):
    filter_type = request.GET.get('filter_type', 'daily')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    orders = get_filtered_orders(filter_type, start_date, end_date).order_by('-ordered_at')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{filter_type}_{timezone.localdate()}.pdf"'
    
    stats = orders.aggregate(
        total_orders_count=Count('id'),
        total_subtotal_sum=Sum('subtotal'),
        total_discount_sum=Sum('discount_amount'),
        total_delivery_sum=Sum('delivery_charge'),
        total_amount_sum=Sum('total_amount')
    )
    
    total_orders_count = stats['total_orders_count'] or 0
    total_subtotal = stats['total_subtotal_sum'] or Decimal('0.00')
    total_discount = stats['total_discount_sum'] or Decimal('0.00')
    total_delivery = stats['total_delivery_sum'] or Decimal('0.00')
    net_revenue = stats['total_amount_sum'] or Decimal('0.00')
    
    # Build Document
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],fontName='Helvetica-Bold',
        fontSize=24,leading=28,
        textColor=colors.HexColor('#111827'), spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'SubtitleStyle', parent=styles['Normal'],fontName='Helvetica',
        fontSize=10,textColor=colors.HexColor('#4B5563'),spaceAfter=25
    )
    header_style = ParagraphStyle(
        'HeaderStyle', parent=styles['Heading3'],
        fontName='Helvetica-Bold',fontSize=12,
        textColor=colors.HexColor('#111827'), spaceAfter=10
    )
    
    story.append(Paragraph("STYLE-HUB SALES REPORT", title_style))
    story.append(Paragraph(f"Generated on {timezone.localtime().strftime('%d %b %Y, %I:%M %p')} | Filter: {filter_type.upper()}", subtitle_style))
    
    summary_data = [
        ['Total Orders', 'Subtotal', 'Coupon Discounts', 'Delivery Charges', 'Net Revenue'],
        [
            str(total_orders_count),
            f"INR {total_subtotal:,.2f}",
            f"INR {total_discount:,.2f}",
            f"INR {total_delivery:,.2f}",
            f"INR {net_revenue:,.2f}"
        ]
    ]
    summary_table = Table(summary_data, colWidths=[100, 110, 110, 110, 120])

    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#374151')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 1), (-1, 1), 11),
        ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#111827')),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#FFFFFF')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 12),
        ('TOPPADDING', (0, 1), (-1, 1), 12),
    ]))
    
    story.append(Paragraph("Executive Performance Summary", header_style))
    story.append(summary_table)
    story.append(Spacer(1, 25))
    
    
    story.append(Paragraph("Detailed Order Breakdown", header_style))
    
    order_headers = ['Order No.', 'Customer', 'Date', 'Payment', 'Status', 'Subtotal', 'Discount', 'Total']
    order_rows = [order_headers]
    
    for order in orders:
        order_rows.append([
            order.order_number,
            order.user.username if order.user else 'Guest',
            timezone.localtime(order.ordered_at).strftime('%d-%m-%Y\n%I:%M %p'),
            order.payment_method,
            order.order_status,
            f"INR {order.subtotal:.2f}",
            f"INR {order.discount_amount:.2f}",
            f"INR {order.total_amount:.2f}"
        ])
        
    orders_table = Table(order_rows, colWidths=[85, 70, 80, 60, 75, 60, 60, 60])
    orders_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),

        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7.5),

        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('ALIGN', (2, 0), (4, -1), 'CENTER'),
        ('ALIGN', (5, 0), (7, -1), 'RIGHT'),

        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),

        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
    ]))
    
    story.append(orders_table)
    doc.build(story)
    return response





@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='admin_login')
def download_sales_report_excel(request):
    filter_type = request.GET.get('filter_type', 'daily')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    orders = get_filtered_orders(filter_type, start_date, end_date).order_by('-ordered_at')
    
    
    stats = orders.aggregate(
        total_orders_count=Count('id'),
        total_subtotal_sum=Sum('subtotal'),
        total_discount_sum=Sum('discount_amount'),
        total_delivery_sum=Sum('delivery_charge'),
        total_amount_sum=Sum('total_amount')
    )
    
    total_orders_count = stats['total_orders_count'] or 0
    total_subtotal = float(stats['total_subtotal_sum'] or 0)
    total_discount = float(stats['total_discount_sum'] or 0)
    total_delivery = float(stats['total_delivery_sum'] or 0)
    net_revenue = float(stats['total_amount_sum'] or 0)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    
    title_font = Font(name="Arial", size=16, bold=True, color="111827")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    section_font = Font(name="Arial", size=11, bold=True, color="1F2937")
    normal_font = Font(name="Arial", size=9, color="374151")
    bold_font = Font(name="Arial", size=9, bold=True, color="000000")
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    summary_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    
    ws.merge_cells("A1:H1")
    ws["A1"] = "STYLE-HUB SALES REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 40
    
    ws["A2"] = f"Report Date: {timezone.now().strftime('%Y-%m-%d %H:%M')} | Filter: {filter_type.upper()}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="4B5563")
    ws["A2"].alignment = align_left
    
    
    ws["A4"] = "Summary Statistics"
    ws["A4"].font = section_font
    
    summary_headers = ["Total Orders", "Subtotal Sales", "Coupon Discounts", "Delivery Charges", "Net Revenue"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = summary_fill
        cell.alignment = align_center
        
    ws.cell(row=6, column=1,value=total_orders_count).alignment=align_center
    ws.cell(row=6, column=2, value=total_subtotal).number_format='$#,##0.00'
    ws.cell(row=6, column=3, value=total_discount).number_format='$#,##0.00'
    ws.cell(row=6, column=4,value=total_delivery).number_format='$#,##0.00'
    ws.cell(row=6, column=5, value=net_revenue).number_format='$#,##0.00'
    
    for c in range(1, 6):
        cell = ws.cell(row=6, column=c)
        cell.font = bold_font
        if c > 1:
            cell.alignment = align_right
            

    ws["A8"] = "Order Details Table"
    ws["A8"].font = section_font
    
    headers = ["Order Number", "Customer", "Order Date", "Payment Method", "Order Status", "Subtotal", "Coupon Discount", "Total Amount"]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_left if col_idx <= 2 else align_center
        
    ws.row_dimensions[9].height = 25
    
    row_idx = 10
    for order in orders:
        ws.cell(row=row_idx, column=1, value=order.order_number).alignment = align_left

        ws.cell(row=row_idx, column=2, value=order.user.username if order.user else 'Guest').alignment = align_left

        ws.cell(row=row_idx, column=3, value=order.ordered_at.strftime('%Y-%m-%d')).alignment = align_center

        ws.cell(row=row_idx, column=4, value=order.payment_method).alignment = align_center

        ws.cell(row=row_idx, column=5, value=order.order_status).alignment = align_center
        
        c6 = ws.cell(row=row_idx, column=6, value=float(order.subtotal))
        c6.number_format = '$#,##0.00'
        c6.alignment = align_right
        
        c7 = ws.cell(row=row_idx, column=7, value=float(order.discount_amount))
        c7.number_format = '$#,##0.00'
        c7.alignment = align_right
        
        c8 = ws.cell(row=row_idx, column=8, value=float(order.total_amount))
        c8.number_format = '$#,##0.00'
        c8.alignment = align_right
        
        for c in range(1, 9):
            ws.cell(row=row_idx, column=c).font = normal_font
            
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1
        
    # Auto-adjust columns width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{filter_type}_{timezone.localdate()}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='admin_login')
def admin_referrals(request):
    search = request.GET.get('search', '')
    referrals = Referral.objects.select_related('referrer', 'referred_user').order_by('-created_at')
    
    if search:
        referrals = referrals.filter(
            Q(referrer__username__icontains=search) |
            Q(referred_user__username__icontains=search) |
            Q(referrer__email__icontains=search) |
            Q(referred_user__email__icontains=search)
        )
        
    paginator = Paginator(referrals, 5)
    page_number = request.GET.get('page')
    referrals_page = paginator.get_page(page_number)
    
    context = {
        'referrals': referrals_page,
        'search': search,
    }
    return render(request, 'admindashboard/referral_management.html', context)


